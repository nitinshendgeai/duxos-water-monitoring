"""Core Excel -> PostgreSQL migration logic.

Shared by the standalone CLI tool (scripts/migrate_excel.py) and the
one-time admin import endpoint, so the validation/import rules live in
exactly one place. See scripts/migrate_excel.py's module docstring for
the full behavior notes (idempotency, skipped sheets, timezone handling).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

from openpyxl.workbook import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import models
from app.config import get_settings

SETTINGS = get_settings()
LOCAL_TZ = ZoneInfo(SETTINGS.timezone)

VALID_SESSIONS = {"Morning", "Afternoon", "Night"}
VALID_ROLES = {"security", "technical", "manager", "gym_attendant"}
VALID_EVENT_TYPES = {"in", "out"}


@dataclass
class SheetReport:
    sheet: str
    imported: int = 0
    skipped_duplicate: int = 0
    failed: list[dict] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "sheet": self.sheet,
            "imported": self.imported,
            "skipped_duplicate": self.skipped_duplicate,
            "failed_count": len(self.failed),
            "failed": self.failed,
        }


def _localize(dt: datetime) -> datetime:
    """Sheets/Excel datetimes come back naive, representing spreadsheet-local
    time (the same timezone Code.gs's Session.getScriptTimeZone() used)."""
    if dt.tzinfo is not None:
        return dt.astimezone(LOCAL_TZ)
    return dt.replace(tzinfo=LOCAL_TZ)


def _parse_iso_utc(value: str) -> datetime:
    """Parses the 'Timestamp' column, a JS .toISOString() UTC string
    (e.g. '2026-08-05T08:03:41.201Z')."""
    return datetime.fromisoformat(value.replace("Z", "+00:00")).astimezone(timezone.utc)


def _clean_str_number(value) -> str:
    """Att_Staff.Phone/Code come back from openpyxl as floats (e.g.
    9930803099.0) since Sheets exported them as plain numbers. Codes are
    always 4-digit and phone numbers never start with 0 in this dataset,
    so int() round-tripping is lossless."""
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value))
    return str(value).strip()


def _row_dicts(ws):
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None and c != "" for c in row):
            continue
        yield dict(zip(headers, row))


def migrate_meter_readings(db: Session, ws, dry_run: bool) -> SheetReport:
    report = SheetReport(sheet="Meter_Readings")
    for i, row in enumerate(_row_dicts(ws), start=2):
        try:
            created_at = _localize(row["Created Timestamp"])
            exists = db.execute(
                select(models.MeterReading.id).where(
                    models.MeterReading.created_at == created_at
                )
            ).first()
            if exists:
                report.skipped_duplicate += 1
                continue

            reading = models.MeterReading(
                reading_date=row["Date"].date(),
                reading_time=row["Time"],
                recorded_by=str(row["Recorded By"]),
                wing_a=float(row["Wing A Meter Reading"]),
                wing_b=float(row["Wing B Meter Reading"]),
                wing_c=float(row["Wing C Meter Reading"]),
                remarks=row.get("Remarks") or None,
                created_at=created_at,
            )
            if not dry_run:
                db.add(reading)
                db.flush()
            report.imported += 1
        except Exception as exc:
            report.failed.append({"row": i, "reason": str(exc)})
    return report


def migrate_tank_status(db: Session, ws, dry_run: bool) -> SheetReport:
    report = SheetReport(sheet="Tank_Status")
    for i, row in enumerate(_row_dicts(ws), start=2):
        try:
            session_name = row["Session"]
            if session_name not in VALID_SESSIONS:
                raise ValueError(f"Unknown session '{session_name}'")

            created_at = _localize(row["Created Timestamp"])
            exists = db.execute(
                select(models.TankStatus.id).where(
                    models.TankStatus.created_at == created_at
                )
            ).first()
            if exists:
                report.skipped_duplicate += 1
                continue

            tank = models.TankStatus(
                reading_date=row["Date"].date(),
                reading_time=row["Time"],
                session=session_name,
                ug_domestic_ab=float(row["UG Domestic Water A+B"]),
                ug_domestic_c=float(row["UG Domestic Water C"]),
                ug_flushing_ab=float(row["UG Flushing Water A+B"]),
                ug_flushing_c=float(row["UG Flushing Water C"]),
                fire_tank=float(row["Fire Tank Level"]),
                oh_dom_a=float(row["OH Domestic A"]),
                oh_dom_b=float(row["OH Domestic B"]),
                oh_dom_c=float(row["OH Domestic C"]),
                oh_flush_a=float(row["OH Flushing A"]),
                oh_flush_b=float(row["OH Flushing B"]),
                oh_flush_c=float(row["OH Flushing C"]),
                remarks=row.get("Remarks") or None,
                created_at=created_at,
            )
            if not dry_run:
                db.add(tank)
                db.flush()
            report.imported += 1
        except Exception as exc:
            report.failed.append({"row": i, "reason": str(exc)})
    return report


def migrate_staff(db: Session, ws, dry_run: bool) -> SheetReport:
    report = SheetReport(sheet="Att_Staff")
    for i, row in enumerate(_row_dicts(ws), start=2):
        try:
            staff_id = str(row["ID"]).strip()
            role = str(row["Role"]).strip()
            if role not in VALID_ROLES:
                raise ValueError(f"Unknown role '{role}'")
            if not staff_id or not row.get("Name"):
                raise ValueError("Missing ID or Name")

            if db.get(models.Staff, staff_id) is not None:
                report.skipped_duplicate += 1
                continue

            staff = models.Staff(
                id=staff_id,
                name=str(row["Name"]).strip(),
                role=role,
                phone=_clean_str_number(row.get("Phone")) or None,
                code=_clean_str_number(row.get("Code")),
            )
            if not dry_run:
                db.add(staff)
                db.flush()
            report.imported += 1
        except Exception as exc:
            report.failed.append({"row": i, "reason": str(exc)})
    return report


def migrate_attendance(
    db: Session, ws, dry_run: bool, known_staff_ids: set[str]
) -> SheetReport:
    report = SheetReport(sheet="Att_Attendance")
    for i, row in enumerate(_row_dicts(ws), start=2):
        try:
            staff_id = str(row["StaffID"]).strip()
            event_type = str(row["Type"]).strip()
            if event_type not in VALID_EVENT_TYPES:
                raise ValueError(f"Unknown type '{event_type}'")
            if staff_id not in known_staff_ids:
                raise ValueError(f"StaffID '{staff_id}' not found in Att_Staff")

            occurred_at = _parse_iso_utc(str(row["Timestamp"]))
            exists = db.execute(
                select(models.Attendance.id).where(
                    models.Attendance.staff_id == staff_id,
                    models.Attendance.occurred_at == occurred_at,
                )
            ).first()
            if exists:
                report.skipped_duplicate += 1
                continue

            event_date = (
                row["Date"].date() if hasattr(row["Date"], "date") else row["Date"]
            )
            # This export predates the Shift column (see the migration
            # report note) -- historical events are imported with shift=NULL.
            shift = row.get("Shift") or None

            event = models.Attendance(
                staff_id=staff_id,
                event_type=event_type,
                location=str(row["Location"]).strip(),
                shift=shift,
                event_date=event_date,
                occurred_at=occurred_at,
            )
            if not dry_run:
                db.add(event)
                db.flush()
            report.imported += 1
        except Exception as exc:
            report.failed.append({"row": i, "reason": str(exc)})
    return report


def run_migration(db: Session, wb: Workbook, dry_run: bool) -> list[SheetReport]:
    """Runs all four sheet migrations in FK-safe order and commits (or
    rolls back, for dry runs) as a single transaction."""
    reports: list[SheetReport] = []

    if "Att_Staff" in wb.sheetnames:
        reports.append(migrate_staff(db, wb["Att_Staff"], dry_run))
        known_ids = {str(r["ID"]).strip() for r in _row_dicts(wb["Att_Staff"])}
    else:
        known_ids = set()

    if "Meter_Readings" in wb.sheetnames:
        reports.append(migrate_meter_readings(db, wb["Meter_Readings"], dry_run))
    if "Tank_Status" in wb.sheetnames:
        reports.append(migrate_tank_status(db, wb["Tank_Status"], dry_run))
    if "Att_Attendance" in wb.sheetnames:
        reports.append(migrate_attendance(db, wb["Att_Attendance"], dry_run, known_ids))

    if dry_run:
        db.rollback()
    else:
        db.commit()

    return reports
